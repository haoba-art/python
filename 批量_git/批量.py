from email.message import EmailMessage
import openpyxl
import smtplib
import sys
wb=openpyxl.load_workbook('email.xlsx')
ws=wb.active
lastcol=ws.max_column
print(lastcol)
num=ws.cell(row=1,column=lastcol).value
num10={}
for r in range(2,ws.max_row+1):
    num00=ws.cell(row=r,column=lastcol).value
    if num00==10:
        member=ws.cell(row=r,column=1).value
        print(member)
        email=ws.cell(row=r,column=2).value
        num10[member]=email
msg=EmailMessage()
email_sender='15801657992@qq.com'
server='smtp.qq.com'
port=465
passwd='' #自己打开邮箱--设置 --账户--开启服务--IMAP/SMTP 复制 授权码
with smtplib.SMTP_SSL(server,port) as server:
    server.login(email_sender,passwd)
    for member,email in num10.items():
        body='Dear %s\n this is test'%member
        msg['From']=email_sender
        msg['to']=email
        msg['Subject']='TEST'
        msg.set_content(body)
        status=server.send_message(msg)
        if status!={}:
            print('%s%s'%(email,status))
        msg.clear()
    server.quit()
