import glob
import openpyxl 
import os 
xlsx_filenames=glob.glob(os.path.join(os.getcwd()+'/','*.xlsx'))
wb_new=openpyxl.Workbook()
for filename in xlsx_filenames:
    ws_new=wb_new.create_sheet()
    wb=openpyxl.load_workbook(filename)
    ws=wb.active
    for row in ws.iter_rows(min_row=1):
            values=[]
            for cell in row:
                values.append(cell.value)
            ws_new.append(values)
            ws_new.title=ws.title      
wb_new.remove(wb_new['Sheet'])
wb_new.save('merge.xlsx')
wb_new.close()
