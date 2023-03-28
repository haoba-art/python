
import openpyxl 
import os 
wb =openpyxl.load_workbook("coil1.xlsx")
ws = wb.active
title=[cell.value for cell in next(ws.iter_rows(min_row=1))]
for row in ws.iter_rows(min_row=2):
    wb_new=openpyxl.Workbook()
    ws_new=wb_new.active
    ws_new.append(title)
    values=[cell.value for cell in row]
    ws_new.append(values)
    ws_new.title='x'+str(values[0])
    wb_new.save(os.getcwd()+'/'+str(values[0])+'.xlsx')

wb_new.close()
