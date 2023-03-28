# =MID(A1,1,FIND(";",A1)-1)
# =MID(A1,FIND(";",A1)+1,20)
import openpyxl
import math

wb = openpyxl.load_workbook(r"demo2.xlsx")
ws = wb.active
lastRow = ws.max_row

for r in range(2, lastRow + 1):
    b = ws.cell(row=r, column=1).value
    ws.cell(row=r, column=2).value = '=MID("%s",1,FIND(";","%s")-1)' % (b, b)
    ws.cell(row=r, column=3).value = '=MID("%s",FIND(";","%s")+1,20)' % (b, b)

wb.save('test2.xlsx')
wb.close()
