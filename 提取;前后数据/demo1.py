# =MID(A1,1,FIND(";",A1)-1)
# =MID(A1,FIND(";",A1)+1,20)
import openpyxl
import math

wb = openpyxl.load_workbook(r"demo1.xlsx")
ws = wb.active
lastRow = ws.max_row
x = {}
for r in range(2, lastRow + 1):
    x[r] = ws.cell(row=r, column=1).value
    index = 0
    for letter in x[r]:
        index += 1
        if letter == ';':
            break
    ws.cell(row=r, column=2).value = float(x[r][:index - 1])
    ws.cell(row=r, column=3).value = float(x[r][index:])
wb.save('test1.xlsx')
wb.close()
