import glob
import openpyxl 
import os 
xlsx_filenames=glob.glob(os.path.join(os.getcwd()+'/','*.xlsx'))
wb_new=openpyxl.Workbook()
ws_new=wb_new.active
ws_new.title='merge'
is_first=True
i=0
for filename in xlsx_filenames:
    wb=openpyxl.load_workbook(filename)
    ws=wb.active
    i+=1
    if is_first:
        for row in ws.iter_rows(min_row=1):
            values=[]
            for cell in row:
                values.append(cell.value)
                #print(type(values))
            ws_new.append(values)
            is_first=False
    else:
        for row in ws.iter_rows(min_row=2):
            values=[]
            for cell in row:
                values.append(cell.value)
            ws_new.append(values)
wb_new.save('merge.xlsx')
wb_new.close()
print(i)
