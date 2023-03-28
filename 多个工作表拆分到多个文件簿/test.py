
import openpyxl 
import os 
wb =openpyxl.load_workbook("coil1.xlsx")
for ws in wb.worksheets:
    wb_new=openpyxl.Workbook()
    ws_new=wb_new.active
    for row_data in ws.iter_rows():
        for row_cell in row_data:
            ws_new[row_cell.coordinate].value=row_cell.value
    wb_new.save(os.getcwd()+'/'+ws.title+'.xlsx')

wb_new.close()
wb.close()
